import json
import os
import time
import uuid
import asyncio
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import openpyxl
import requests
from openai import OpenAI
import httpx
from google import genai as google_genai

from glossary import get_glossary

load_dotenv()

app = FastAPI(title="Škoda X Translator")

# Statické soubory (fonty)
FONTS_DIR = Path("fonts")
if FONTS_DIR.exists():
    app.mount("/fonts", StaticFiles(directory="fonts"), name="fonts")

UPLOAD_DIR = Path("uploads")
OUTPUT_DIR = Path("outputs")
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# Explicitní httpx klient bez proxy – systémová ALL_PROXY (socks5h) není httpx kompatibilní
client = OpenAI(
    api_key=os.getenv("OPENAI_API_KEY"),
    http_client=httpx.Client(proxy=None),
)

# Gemini klient (fallback za OpenAI)
_gemini_key = os.getenv("GEMINI_API_KEY")
gemini_client = google_genai.Client(api_key=_gemini_key) if _gemini_key else None
GEMINI_MODEL = "gemini-2.0-flash"

# ── Translation tables ────────────────────────────────────────────────────────

language_mapping = {
    "bg-bg": "BG", "cs-cz": "CS", "da-dk": "DA",
    "de-at": "DE", "de-ch": "DE", "de-de": "DE", "de-lu": "DE",
    "el-gr": "EL", "en-gb": "EN-GB", "en-ie": "EN-GB",
    "es-es": "ES", "es-ic": "ES", "et-ee": "ET", "fi-fi": "FI",
    "fr-be": "FR", "fr-ch": "FR", "fr-fr": "FR", "fr-lu": "FR",
    "hu-hu": "HU", "it-ch": "IT", "it-it": "IT",
    "lt-lt": "LT", "lv-lv": "LV",
    "nl-be": "NL", "nl-nl": "NL", "nn-no": "NB",
    "pl-pl": "PL", "pt-pt": "PT-PT", "ro-ro": "RO",
    "sk-sk": "SK", "sl-si": "SL", "sv-se": "SV", "uk-ua": "UK",
}

chatgpt_model = "gpt-4o-mini"
chatgpt_languages = {
    "bs-ba": "Bosnian", "hr-hr": "Croatian",
    "mk-mk": "Macedonian", "sr-rs": "Serbian",
    "is-is": "Icelandic",
}
chatgpt_informal_languages = {"nl-nl": chatgpt_model}
chatgpt_skoda_languages = {lang: chatgpt_model for lang in language_mapping}

# Mapování DeepL kódů na názvy jazyků pro OpenAI fallback
deepl_to_lang_name = {
    "BG": "Bulgarian", "CS": "Czech", "DA": "Danish", "DE": "German",
    "EL": "Greek", "EN-GB": "English (British)", "ES": "Spanish",
    "ET": "Estonian", "FI": "Finnish", "FR": "French", "HU": "Hungarian",
    "IT": "Italian", "LT": "Lithuanian", "LV": "Latvian", "NL": "Dutch",
    "NB": "Norwegian", "PL": "Polish", "PT-PT": "Portuguese",
    "RO": "Romanian", "SK": "Slovak", "SL": "Slovenian", "SV": "Swedish",
    "UK": "Ukrainian",
}

TRANSLATION_MODE_DEEPL = "deepl"
TRANSLATION_MODE_OPENAI = "openai"
TRANSLATION_MODE_OPENAI_SKODA = "openai-skoda"

# ── In-memory job store ───────────────────────────────────────────────────────

jobs: dict[str, dict] = {}

# Global flag: pokud DeepL vrátí 456 (quota), přepneme na OpenAI pro celý zbytek jobu
deepl_quota_exceeded = False


def set_progress(job_id: str, message: str, percent: int):
    if job_id in jobs:
        jobs[job_id]["progress"].append({"message": message, "percent": percent})
        jobs[job_id]["percent"] = percent


# ── Translation helpers ───────────────────────────────────────────────────────

class DeepLQuotaExceeded(Exception):
    pass


def call_deepl(msgs: list, trg_lang: str) -> list:
    request_body = {"text": msgs, "source_lang": "EN", "target_lang": trg_lang}
    for attempt in range(3):
        r = requests.post(
            "https://api-free.deepl.com/v2/translate",
            headers={
                "Authorization": f"DeepL-Auth-Key {os.getenv('DEEPL_API_KEY')}",
                "Content-Type": "application/json",
            },
            data=json.dumps(request_body).encode(),
        )
        if r.status_code == 456:
            # 456 = DeepL quota exceeded
            raise DeepLQuotaExceeded("DeepL quota vyčerpána")
        if r.status_code in (429, 503):
            time.sleep(5)
            continue
        if r.status_code != 200:
            raise HTTPException(status_code=502, detail=f"DeepL error {r.status_code}: {r.text}")
        return [obj["text"] for obj in r.json()["translations"]]
    # Po 3 opakovaných 429/503 taky přepneme na OpenAI
    raise DeepLQuotaExceeded("DeepL opakovaně nedostupný – přepínám na OpenAI")


def call_openai_skoda(msg: str, lang_code: str) -> str:
    if msg.startswith("https://") or "@" in msg:
        return msg
    glossary = get_glossary(lang_code)
    if not glossary:
        return msg
    resp = client.chat.completions.create(
        model=chatgpt_model,
        messages=[
            {"role": "system", "content": (
                "You are a translation corrector for Škoda X. "
                "Your task is ONLY to correct terminology according to the mandatory Škoda X "
                "terminology listed below. Do not correct grammar or style. "
                "Return only the corrected text, without any comments.\n\n" + glossary
            )},
            {"role": "user", "content": msg},
        ],
    )
    return resp.choices[0].message.content


def call_openai_dutch_informal(msg: str) -> str:
    if msg.startswith("https://") or "@" in msg:
        return msg
    resp = client.chat.completions.create(
        model=chatgpt_model,
        messages=[
            {"role": "system", "content": (
                "Change dutch text, which is polite Tone Of Voice (U/Uw) to informal TOV (je/jij/jouw). "
                "Don't change the style of the text, and don't add any numbers."
            )},
            {"role": "user", "content": msg},
        ],
    )
    return resp.choices[0].message.content


def call_openai(msg: str, trg_lang: str, lang_code: str = "") -> str:
    if msg.startswith("https://") or "@" in msg:
        return msg
    glossary = get_glossary(lang_code) if lang_code else ""
    glossary_section = (
        "\n\nAlso apply this mandatory Škoda X terminology:\n" + glossary if glossary else ""
    )
    resp = client.chat.completions.create(
        model=chatgpt_model,
        messages=[
            {"role": "system", "content": (
                f"Translate the given text from English to {trg_lang}. "
                f"Don't change the style of the text, and don't add any numbers.{glossary_section}"
            )},
            {"role": "user", "content": msg},
        ],
    )
    return resp.choices[0].message.content


# ── Gemini fallback helpers ───────────────────────────────────────────────────

def _gemini_ask(prompt_system: str, prompt_user: str) -> str:
    if not gemini_client:
        raise RuntimeError("GEMINI_API_KEY není nastaven")
    response = gemini_client.models.generate_content(
        model=GEMINI_MODEL,
        contents=f"{prompt_system}\n\n{prompt_user}",
    )
    return response.text.strip()


def call_gemini(msg: str, trg_lang: str, lang_code: str = "") -> str:
    if msg.startswith("https://") or "@" in msg:
        return msg
    glossary = get_glossary(lang_code) if lang_code else ""
    glossary_section = (
        "\n\nAlso apply this mandatory Škoda X terminology:\n" + glossary if glossary else ""
    )
    system = (
        f"Translate the given text from English to {trg_lang}. "
        f"Return only the translated text, nothing else.{glossary_section}"
    )
    return _gemini_ask(system, msg)


def call_gemini_skoda(msg: str, lang_code: str) -> str:
    if msg.startswith("https://") or "@" in msg:
        return msg
    glossary = get_glossary(lang_code)
    if not glossary:
        return msg
    system = (
        "You are a translation corrector for Škoda X. "
        "ONLY correct terminology according to the mandatory Škoda X terminology below. "
        "Do not change grammar or style. Return only the corrected text.\n\n" + glossary
    )
    return _gemini_ask(system, msg)


def call_gemini_dutch_informal(msg: str) -> str:
    if msg.startswith("https://") or "@" in msg:
        return msg
    system = (
        "Change dutch text from polite Tone Of Voice (U/Uw) to informal TOV (je/jij/jouw). "
        "Don't change style, don't add numbers. Return only the changed text."
    )
    return _gemini_ask(system, msg)


def call_ai(msg: str, trg_lang: str, lang_code: str = "") -> str:
    """Gemini jako první, OpenAI jako záloha."""
    try:
        return call_gemini(msg, trg_lang, lang_code)
    except Exception:
        return call_openai(msg, trg_lang, lang_code)


def call_ai_skoda(msg: str, lang_code: str) -> str:
    try:
        return call_gemini_skoda(msg, lang_code)
    except Exception:
        return call_openai_skoda(msg, lang_code)


def call_ai_dutch_informal(msg: str) -> str:
    try:
        return call_gemini_dutch_informal(msg)
    except Exception:
        return call_openai_dutch_informal(msg)


# ── Core translation job ──────────────────────────────────────────────────────

def run_translation(job_id: str, input_path: Path, output_path: Path):
    try:
        jobs[job_id]["status"] = "running"
        set_progress(job_id, "Načítám soubor…", 2)

        wb_obj = openpyxl.load_workbook(input_path)
        sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row

        wb_write = openpyxl.Workbook()
        sheet_write = wb_write.active

        target_langs = []
        source_msgs = {}
        translation_cache = {}

        for row in range(1, max_row + 1):
            key_name = None
            for col in range(1, max_col + 1):
                cell = sheet_obj.cell(row, col)
                sheet_write.cell(row, col).value = cell.value
                if row == 1:
                    if col > 2 and cell.value and "-" in str(cell.value):
                        target_langs.append(cell.value)
                else:
                    if col == 1:
                        key_name = f"{cell.value}_{row}"
                    elif col == 2 and cell.value is not None:
                        source_msgs[key_name] = cell.value
                    else:
                        break

        total_langs = len(target_langs)
        set_progress(job_id, f"Nalezeno {len(source_msgs)} textů, {total_langs} jazyků", 5)

        target_col = 2
        deepl_quota_exceeded = False  # reset per job

        for lang_idx, lang in enumerate(target_langs, 1):
            target_col += 1
            lang_lower = lang.lower()
            lang_code = lang_lower.split("-")[0]

            percent_start = 5 + int((lang_idx - 1) / total_langs * 90)
            percent_end = 5 + int(lang_idx / total_langs * 90)

            if lang_lower in language_mapping:
                target_lang = language_mapping[lang_lower]
                if lang_lower in chatgpt_skoda_languages:
                    mode = TRANSLATION_MODE_OPENAI_SKODA
                else:
                    mode = TRANSLATION_MODE_DEEPL
            elif lang_lower in chatgpt_languages:
                target_lang = chatgpt_languages[lang_lower]
                mode = TRANSLATION_MODE_OPENAI
            else:
                raise ValueError(f"Nepodporovaný jazyk: {lang}")

            set_progress(job_id, f"[{lang_idx}/{total_langs}] {lang} – překládám…", percent_start)

            # Cache key zahrnuje i "informal" aby nl-nl a nl-be nesdílely cache
            cache_key = f"{target_lang}_informal" if lang_lower in chatgpt_informal_languages else target_lang

            if cache_key in translation_cache:
                set_progress(job_id, f"[{lang_idx}/{total_langs}] {lang} – z cache ✓", percent_end)
                r_nr = 2
                for key_name in source_msgs:
                    sheet_write.cell(r_nr, target_col).value = translation_cache[cache_key].get(key_name)
                    r_nr += 1
                continue

            # Název jazyka pro OpenAI fallback (místo DeepL kódu "NL", "PT-PT" atd.)
            fallback_lang_name = deepl_to_lang_name.get(target_lang, target_lang)

            if mode in (TRANSLATION_MODE_DEEPL, TRANSLATION_MODE_OPENAI_SKODA):
                if deepl_quota_exceeded:
                    # DeepL quota already hit – přeložit přes Gemini/OpenAI
                    set_progress(job_id, f"[{lang_idx}/{total_langs}] {lang} – AI překlad (DeepL quota)…", percent_start)
                    translated = [call_ai(msg, fallback_lang_name, lang_code) for msg in source_msgs.values()]
                    # Neformální tón pro nl-nl i v fallbacku
                    if lang_lower in chatgpt_informal_languages:
                        set_progress(job_id, f"[{lang_idx}/{total_langs}] {lang} – neformální tón…", percent_start + 1)
                        translated = [call_ai_dutch_informal(t) for t in translated]
                else:
                    try:
                        translated = call_deepl(list(source_msgs.values()), target_lang)
                    except DeepLQuotaExceeded as e:
                        deepl_quota_exceeded = True
                        set_progress(job_id, f"⚠ {e} – přepínám na Gemini/OpenAI pro zbytek překladu", percent_start)
                        translated = [call_ai(msg, fallback_lang_name, lang_code) for msg in source_msgs.values()]
                        if lang_lower in chatgpt_informal_languages:
                            set_progress(job_id, f"[{lang_idx}/{total_langs}] {lang} – neformální tón…", percent_start + 1)
                            translated = [call_ai_dutch_informal(t) for t in translated]

                    if not deepl_quota_exceeded:
                        if lang_lower in chatgpt_informal_languages:
                            set_progress(job_id, f"[{lang_idx}/{total_langs}] {lang} – neformální tón…", percent_start + 1)
                            translated = [call_ai_dutch_informal(t) for t in translated]

                        if mode == TRANSLATION_MODE_OPENAI_SKODA:
                            set_progress(job_id, f"[{lang_idx}/{total_langs}] {lang} – Škoda terminologie…", percent_start + 2)
                            translated = [call_ai_skoda(t, lang_code) for t in translated]

            elif mode == TRANSLATION_MODE_OPENAI:
                translated = []
                for source_msg in source_msgs.values():
                    translated.append(call_ai(source_msg, target_lang, lang_code))

            translation_cache[cache_key] = {}
            r_nr = 2
            for i, key_name in enumerate(source_msgs):
                sheet_write.cell(r_nr, target_col).value = translated[i]
                translation_cache[cache_key][key_name] = translated[i]
                r_nr += 1

            set_progress(job_id, f"[{lang_idx}/{total_langs}] {lang} ✓", percent_end)

        wb_write.save(output_path)
        wb_obj.close()

        set_progress(job_id, "Hotovo!", 100)
        jobs[job_id]["status"] = "done"

    except Exception as e:
        jobs[job_id]["status"] = "error"
        jobs[job_id]["error"] = str(e)


# ── API endpoints ─────────────────────────────────────────────────────────────

@app.post("/translate")
async def start_translation(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Pouze .xlsx soubory")

    job_id = str(uuid.uuid4())
    input_path = UPLOAD_DIR / f"{job_id}_input.xlsx"
    output_path = OUTPUT_DIR / f"{job_id}_output.xlsx"

    content = await file.read()
    with open(input_path, "wb") as f:
        f.write(content)

    jobs[job_id] = {
        "status": "queued",
        "progress": [],
        "percent": 0,
        "output": str(output_path),
        "filename": file.filename.replace(".xlsx", "_translated.xlsx"),
    }

    background_tasks.add_task(run_translation, job_id, input_path, output_path)
    return {"job_id": job_id}


@app.get("/status/{job_id}")
def get_status(job_id: str):
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job nenalezen")
    job = jobs[job_id]
    return {
        "status": job["status"],
        "percent": job["percent"],
        "messages": [p["message"] for p in job["progress"]],
        "error": job.get("error"),
    }


@app.get("/download/{job_id}")
def download_result(job_id: str):
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job nenalezen")
    job = jobs[job_id]
    if job["status"] != "done":
        raise HTTPException(status_code=400, detail="Překlad ještě není dokončen")
    return FileResponse(
        path=job["output"],
        filename=job["filename"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/template")
def download_template():
    import io
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    headers = [
        "key_name", "English",
        "bg-BG", "bs-BA", "cs-CZ", "da-DK",
        "de-AT", "de-CH", "de-DE", "de-LU",
        "el-GR", "en-GB", "en-IE",
        "es-ES", "es-IC", "et-EE", "fi-FI",
        "fr-BE", "fr-CH", "fr-FR", "fr-LU",
        "hr-HR", "hu-HU", "is-IS",
        "it-CH", "it-IT", "lt-LT", "lv-LV",
        "mk-MK", "nl-BE", "nl-NL", "nn-NO",
        "pl-PL", "pt-PT", "ro-RO",
        "sk-SK", "sl-SI", "sr-RS", "sv-SE", "uk-UA",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Translations"

    dark_fill  = PatternFill(start_color="0E3A2F", end_color="0E3A2F", fill_type="solid")
    key_fill   = PatternFill(start_color="1a5c47", end_color="1a5c47", fill_type="solid")
    green_font = Font(color="78FAAE", bold=True, name="Calibri", size=11)
    white_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    center     = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, h)
        cell.font      = green_font
        cell.alignment = center
        cell.fill      = dark_fill
        if h == "key_name":
            cell.fill = key_fill
            cell.font = white_font

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 48
    for i in range(3, len(headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = 14
    ws.row_dimensions[1].height = 24

    # Example row with naming convention
    gray_fill   = PatternFill(start_color="F0F4F2", end_color="F0F4F2", fill_type="solid")
    italic_font = Font(color="888888", italic=True, name="Calibri", size=10)
    example_key = "YYMM_Product_CampaignName_AuthorInitials_XX_Title"
    example_en  = "Your source English text goes here"
    ws.cell(2, 1, example_key).font  = italic_font
    ws.cell(2, 2, example_en).font   = italic_font
    for col_idx in range(1, len(headers) + 1):
        ws.cell(2, col_idx).fill = gray_fill
    ws.row_dimensions[2].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=skoda_x_translation_template.xlsx"},
    )


@app.get("/", response_class=HTMLResponse)
def index():
    html_path = Path("templates/index.html")
    if html_path.exists():
        return HTMLResponse(html_path.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>Škoda X Translator</h1>")
